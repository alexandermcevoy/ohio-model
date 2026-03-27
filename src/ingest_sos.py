"""
ingest_sos.py — Parse Ohio Secretary of State precinct-level XLSX files.

All SOS general election XLSX files share the same 2-row header layout:
  Row 1 : race names (merged cells — only first column of each race has a value)
  Row 2 : candidate names, e.g. "Nan Whaley and Cheryl L. Stephens (D)"
  Row 3 : "Total" aggregate row  (skipped)
  Row 4 : "Percentage" row        (skipped)
  Row 5+: one row per precinct

Sheet names vary by year; we detect them by keyword substring.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Statewide partisan races we want: keyword in race cell -> short label.
STATEWIDE_OFFICE_MAP: dict[str, str] = {
    "president": "pre",
    "governor": "gov",
    "u.s. senator": "uss",
    "attorney general": "atg",
    "auditor": "aud",
    "secretary of state": "sos_off",
    "treasurer": "tre",
}

COUNTY_COL = "county_name"
PREC_CODE_COL = "precinct_code"
PREC_NAME_COL = "precinct_name"
META_COLS = 8   # columns before vote data begins

_HOUSE_RE = re.compile(r"State Representative.*?District\s+(\d+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class RaceSpec:
    label: str        # short label, e.g. 'gov'
    office: str       # full cleaned office name
    d_cols: list[str] # DataFrame column names for D candidates
    r_cols: list[str] # DataFrame column names for R candidates
    d_candidate_names: list[str] = field(default_factory=list)  # raw name strings from XLSX header
    r_candidate_names: list[str] = field(default_factory=list)  # raw name strings from XLSX header

    def has_contest(self) -> bool:
        return bool(self.d_cols and self.r_cols)


@dataclass
class SosFile:
    year: str
    path: Path
    statewide: dict[str, RaceSpec]    # label -> RaceSpec
    house: dict[int, RaceSpec]        # district_num -> RaceSpec
    precinct_statewide: pd.DataFrame  # county_name, precinct_code, [vote cols]
    precinct_house: pd.DataFrame      # county_name, precinct_code, [house vote cols]


# ---------------------------------------------------------------------------
# Header parsing
# ---------------------------------------------------------------------------

def _extract_party(name: str) -> str:
    s = name.strip()
    if "(WI)*" in s:
        return "WI"
    m = re.search(r"\(([A-Z])\)\s*$", s)
    return m.group(1) if m else "O"


def _parse_sheet(ws, mode: str) -> tuple[dict, pd.DataFrame]:
    """
    Parse one SOS worksheet.

    Parameters
    ----------
    ws : openpyxl worksheet (read-only)
    mode : 'statewide' or 'house'

    Returns
    -------
    race_map : dict
        For 'statewide': {label: RaceSpec}
        For 'house':     {district_num: RaceSpec}
    df : DataFrame
        Precinct rows with columns [county_name, precinct_code, precinct_name,
        <vote_col_1>, <vote_col_2>, ...]
    """
    # ── Read the 2 header rows ───────────────────────────────────────────────
    header_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    row0 = list(header_rows[0])
    row1 = list(header_rows[1])
    n = max(len(row0), len(row1))
    row0 += [None] * (n - len(row0))
    row1 += [None] * (n - len(row1))

    # Forward-fill race names across merged cells
    col_race: list[str | None] = []
    cur = None
    for v in row0:
        if v is not None:
            cur = re.sub(r"\s+", " ", str(v)).strip()
        col_race.append(cur)

    # ── Map each column to its race and party ────────────────────────────────
    # race_key -> {'d': [col_indices], 'r': [col_indices]}
    race_cols: dict[object, dict[str, list[int]]] = {}

    for i in range(META_COLS, n):
        race_raw = col_race[i] if i < len(col_race) else None
        cand = row1[i] if i < len(row1) else None
        if not race_raw or not cand:
            continue

        party = _extract_party(str(cand))
        if party not in ("D", "R"):
            continue

        if mode == "statewide":
            lower = race_raw.lower()
            label = next(
                (lbl for kw, lbl in STATEWIDE_OFFICE_MAP.items() if kw in lower),
                None,
            )
            if label is None:
                continue
            key = label
        else:  # house
            m = _HOUSE_RE.search(race_raw)
            if not m:
                continue
            key = int(m.group(1))

        race_cols.setdefault(key, {"d": [], "r": [], "d_cand": [], "r_cand": []})
        party_key = "d" if party == "D" else "r"
        race_cols[key][party_key].append(i)
        race_cols[key][f"{party_key}_cand"].append(str(cand).strip())

    # ── Build column name lists and RaceSpec objects ─────────────────────────
    vote_indices: list[int] = []
    vote_names: list[str] = []
    race_map: dict = {}

    for key, dr in race_cols.items():
        label = str(key) if mode == "house" else key
        d_names, r_names = [], []

        for j, idx in enumerate(dr["d"]):
            nm = f"{key}_d{j}" if mode == "house" else f"{label}_d{j}"
            vote_indices.append(idx)
            vote_names.append(nm)
            d_names.append(nm)

        for j, idx in enumerate(dr["r"]):
            nm = f"{key}_r{j}" if mode == "house" else f"{label}_r{j}"
            vote_indices.append(idx)
            vote_names.append(nm)
            r_names.append(nm)

        office_raw = col_race[dr["d"][0] if dr["d"] else dr["r"][0]]
        race_map[key] = RaceSpec(
            label=str(key) if mode == "house" else label,
            office=office_raw or "",
            d_cols=d_names,
            r_cols=r_names,
            d_candidate_names=dr.get("d_cand", []),
            r_candidate_names=dr.get("r_cand", []),
        )

    # ── Read precinct rows ───────────────────────────────────────────────────
    id_indices = [0, 1, 2]
    id_names = [COUNTY_COL, PREC_NAME_COL, PREC_CODE_COL]
    all_indices = id_indices + vote_indices
    all_names = id_names + vote_names

    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        v0 = str(row[0]).strip()
        if v0 in ("Total", "Percentage", ""):
            continue
        record = {}
        for idx, nm in zip(all_indices, all_names):
            record[nm] = row[idx] if idx < len(row) else None
        records.append(record)

    df = pd.DataFrame(records)
    for c in [COUNTY_COL, PREC_NAME_COL, PREC_CODE_COL]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.upper()
    for nm in vote_names:
        if nm in df.columns:
            df[nm] = pd.to_numeric(df[nm], errors="coerce").fillna(0.0)

    return race_map, df


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def _find_sheet(wb, *keywords: str):
    for name in wb.sheetnames:
        lower = name.lower()
        if any(kw.lower() in lower for kw in keywords):
            return wb[name]
    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_sos_file(path: str | Path) -> SosFile:
    """Load and parse an Ohio SOS general election precinct XLSX."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"SOS file not found: {path}")

    print(f"\nLoading SOS file: {path.name} …")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Detect year from first cell of first sheet
    first_row = next(wb[wb.sheetnames[0]].iter_rows(max_row=1, values_only=True))
    ym = re.search(r"(201[0-9]|202[0-9])", str(first_row[0]))
    year = ym.group(1) if ym else "unknown"
    print(f"  Detected year: {year}")

    id_cols = [COUNTY_COL, PREC_NAME_COL, PREC_CODE_COL]

    # ── Statewide races ──────────────────────────────────────────────────────
    sw_race_map: dict[str, RaceSpec] = {}
    sw_dfs: list[pd.DataFrame] = []

    # Presidential sheet
    ws = _find_sheet(wb, "president")
    if ws:
        rm, df = _parse_sheet(ws, "statewide")
        sw_race_map.update(rm)
        sw_dfs.append(df)

    # Statewide Offices sheet (Gov, AG, Auditor, SOS, Treasurer)
    ws = _find_sheet(wb, "statewide offices")
    if ws:
        rm, df = _parse_sheet(ws, "statewide")
        sw_race_map.update(rm)
        sw_dfs.append(df)

    # U.S. Congress sheet — extract Senate only
    ws = _find_sheet(wb, "u.s. congress")
    if ws:
        rm, df = _parse_sheet(ws, "statewide")
        if "uss" in rm:
            uss_spec = rm["uss"]
            sw_race_map["uss"] = uss_spec
            # Keep only id cols + uss vote cols
            uss_vote_cols = uss_spec.d_cols + uss_spec.r_cols
            sw_dfs.append(df[id_cols + [c for c in uss_vote_cols if c in df.columns]])

    # Merge all statewide DFs on precinct identifiers (outer join to keep all precincts)
    if sw_dfs:
        precinct_sw = sw_dfs[0]
        for part in sw_dfs[1:]:
            new_cols = [c for c in part.columns if c not in id_cols]
            if new_cols:
                precinct_sw = precinct_sw.merge(
                    part[id_cols + new_cols], on=id_cols, how="outer"
                )
    else:
        precinct_sw = pd.DataFrame(columns=id_cols)

    # ── House races ──────────────────────────────────────────────────────────
    ws = _find_sheet(wb, "general assembly", "gen assembly", "ohio general assembly")
    if ws:
        house_map, precinct_house = _parse_sheet(ws, "house")
    else:
        house_map, precinct_house = {}, pd.DataFrame(columns=id_cols)

    # ── Summary ──────────────────────────────────────────────────────────────
    contested = [k for k, v in sw_race_map.items() if v.has_contest()]
    print(f"  Statewide contested races found: {sorted(contested)}")
    print(f"  House districts parsed: {len(house_map)}")
    _print_totals(precinct_sw, sw_race_map)

    return SosFile(
        year=year,
        path=path,
        statewide=sw_race_map,
        house=house_map,
        precinct_statewide=precinct_sw,
        precinct_house=precinct_house,
    )


def _print_totals(df: pd.DataFrame, race_map: dict[str, RaceSpec]) -> None:
    if df.empty:
        return
    print("  Statewide vote totals:")
    for label in sorted(race_map):
        spec = race_map[label]
        if not spec.has_contest():
            continue
        d = sum(df[c].sum() for c in spec.d_cols if c in df.columns)
        r = sum(df[c].sum() for c in spec.r_cols if c in df.columns)
        tp = d + r
        share = d / tp if tp > 0 else float("nan")
        print(f"    {label:10s}  D={d:>10,.0f}  R={r:>10,.0f}  D-2P={share:.4f}")


def get_race_df(sos: SosFile, label: str) -> pd.DataFrame:
    """Return tidy precinct-level DF: county_name, precinct_code, d_votes, r_votes."""
    spec = sos.statewide.get(label)
    if spec is None:
        raise KeyError(
            f"Race '{label}' not found in {sos.year}. "
            f"Available: {sorted(sos.statewide.keys())}"
        )
    df = sos.precinct_statewide
    d = df[spec.d_cols].sum(axis=1) if spec.d_cols else pd.Series(0.0, index=df.index)
    r = df[spec.r_cols].sum(axis=1) if spec.r_cols else pd.Series(0.0, index=df.index)
    out = df[[COUNTY_COL, PREC_CODE_COL]].copy()
    out["d_votes"] = d.values
    out["r_votes"] = r.values
    return out.dropna(subset=[COUNTY_COL, PREC_CODE_COL]).reset_index(drop=True)
