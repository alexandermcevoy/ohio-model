"""
ingest_historical.py — Parse Ohio SOS 2010–2014 precinct-level XLSX files
for the Census Block Backbone.

These older files have different formats from the 2016+ parser in ingest_sos.py.
This module extracts county-level D/R vote totals for statewide races, which is
what the backbone needs for county → block disaggregation.

Supported files:
  2010precinct.xlsx            — single "AllCounties" sheet, flat layout, no party tags
  2012statewidebyprecinct.xlsx — multi-sheet, party tags in headers
  2014.xlsx                    — single sheet, party tags in headers

For 2016+ files, use ingest_sos.py + aggregate to county level.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# 2010 — Flat "AllCounties" sheet, "Race - CandidateName" headers, no party tags
# ---------------------------------------------------------------------------

# Known D/R candidates for Ohio 2010 statewide races.
# Source: Ohio SOS official results + Ballotpedia.
_2010_STATEWIDE_CANDIDATES: dict[str, dict[str, list[str]]] = {
    "gov": {
        "d": ["Strickland"],
        "r": ["Kasich"],
    },
    "atg": {
        "d": ["Cordray"],
        "r": ["DeWine"],
    },
    "aud": {
        "d": ["Pepper"],   # David Pepper ran for Auditor in 2010
        "r": ["Yost"],
    },
    "sos_off": {
        "d": ["O'Shaughnessy"],
        "r": ["Husted"],
    },
    "tre": {
        "d": ["Boyce"],
        "r": ["Mandel"],
    },
    "uss": {
        "d": ["Fisher"],
        "r": ["Portman"],
    },
}

# Map office keyword (in header) → race label
_2010_OFFICE_MAP: dict[str, str] = {
    "governor": "gov",
    "attorney general": "atg",
    "auditor of state": "aud",
    "secretary of state": "sos_off",
    "treasurer of state": "tre",
    "u.s. senate": "uss",
}


def parse_2010_county_votes(path: str | Path) -> pd.DataFrame:
    """
    Parse 2010 SOS XLSX and return county-level D/R totals per statewide race.

    Returns DataFrame: county_name, race, d_votes, r_votes
    """
    path = Path(path)
    print(f"  Parsing 2010 SOS file: {path.name} …")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["AllCounties"]

    # Row 1 (0-indexed) has headers: "Race - CandidateName"
    header_row = list(list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0])

    # Map column indices to (race_label, party)
    col_mapping: list[tuple[str, str] | None] = [None] * len(header_row)

    for i, h in enumerate(header_row):
        if h is None or i < 6:  # first 6 are metadata
            continue
        h_str = str(h)

        # Determine which statewide race this belongs to
        race_label = None
        for kw, lbl in _2010_OFFICE_MAP.items():
            if kw in h_str.lower():
                race_label = lbl
                break

        if race_label is None:
            continue

        # Determine party by candidate name lookup
        candidates = _2010_STATEWIDE_CANDIDATES.get(race_label, {})
        party = None
        for p, names in candidates.items():
            if any(n.lower() in h_str.lower() for n in names):
                party = p
                break

        if party:
            col_mapping[i] = (race_label, party)

    # Read data rows (start at row 3, 1-indexed row 3 = index 2)
    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] is None:
            continue
        county_name = str(row[1]).strip().upper()
        if county_name in ("", "COUNTY NAME"):
            continue

        for i, mapping in enumerate(col_mapping):
            if mapping is None or i >= len(row):
                continue
            race_label, party = mapping
            votes = row[i] if row[i] is not None else 0
            try:
                votes = float(votes)
            except (ValueError, TypeError):
                votes = 0
            records.append({
                "county_name": county_name,
                "race": race_label,
                "party": party,
                "votes": votes,
            })

    wb.close()

    df = pd.DataFrame(records)

    # Pivot to county_name, race, d_votes, r_votes
    agg = df.groupby(["county_name", "race", "party"])["votes"].sum().reset_index()
    d_votes = agg[agg["party"] == "d"].rename(columns={"votes": "d_votes"}).drop(columns=["party"])
    r_votes = agg[agg["party"] == "r"].rename(columns={"votes": "r_votes"}).drop(columns=["party"])
    result = d_votes.merge(r_votes, on=["county_name", "race"], how="outer").fillna(0)

    _print_summary(result, "2010")
    return result


# ---------------------------------------------------------------------------
# 2012 — Multi-sheet, party tags "(D)"/"(R)" in candidate headers
# ---------------------------------------------------------------------------

_2012_STATEWIDE_OFFICE_MAP: dict[str, str] = {
    "president": "pre",
    "u.s. senator": "uss",
    "u.s. senate": "uss",
}


def parse_2012_county_votes(path: str | Path) -> pd.DataFrame:
    """
    Parse 2012 SOS XLSX. Only President + USS available (no statewide offices).

    Returns DataFrame: county_name, race, d_votes, r_votes
    """
    path = Path(path)
    print(f"  Parsing 2012 SOS file: {path.name} …")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    all_results = []

    # President sheet
    if "President" in wb.sheetnames:
        results = _parse_2012_2014_sheet(wb["President"], _2012_STATEWIDE_OFFICE_MAP, meta_cols=8, data_start_row=5)
        all_results.append(results)

    # U.S. Congress sheet (extract USS)
    if "U.S. Congress" in wb.sheetnames:
        results = _parse_2012_2014_sheet(wb["U.S. Congress"], _2012_STATEWIDE_OFFICE_MAP, meta_cols=8, data_start_row=5)
        all_results.append(results)

    wb.close()

    if not all_results:
        return pd.DataFrame(columns=["county_name", "race", "d_votes", "r_votes"])

    result = pd.concat(all_results, ignore_index=True)
    # Aggregate by county and race (in case of duplicates across sheets)
    result = result.groupby(["county_name", "race"])[["d_votes", "r_votes"]].sum().reset_index()

    _print_summary(result, "2012")
    return result


# ---------------------------------------------------------------------------
# 2014 — Single sheet, party tags "(D)"/"(R)" in candidate headers
# ---------------------------------------------------------------------------

_2014_STATEWIDE_OFFICE_MAP: dict[str, str] = {
    "governor": "gov",
    "attorney general": "atg",
    "auditor of state": "aud",
    "secretary of state": "sos_off",
    "treasurer of state": "tre",
    "u.s. senate": "uss",
    "u.s. senator": "uss",
}


def parse_2014_county_votes(path: str | Path) -> pd.DataFrame:
    """
    Parse 2014 SOS XLSX and return county-level D/R totals per statewide race.

    Returns DataFrame: county_name, race, d_votes, r_votes
    """
    path = Path(path)
    print(f"  Parsing 2014 SOS file: {path.name} …")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 2014: 7 metadata cols, data starts at row 3 (after Total row)
    result = _parse_2012_2014_sheet(ws, _2014_STATEWIDE_OFFICE_MAP, meta_cols=7, data_start_row=3)

    wb.close()
    _print_summary(result, "2014")
    return result


# ---------------------------------------------------------------------------
# Shared parser for 2012/2014 format (2-row header with race + candidate names, party tags)
# ---------------------------------------------------------------------------

def _parse_2012_2014_sheet(
    ws,
    office_map: dict[str, str],
    meta_cols: int,
    data_start_row: int,
) -> pd.DataFrame:
    """
    Parse a sheet with the 2-row header format used by 2012 and 2014 files.

    Row 0: race names (merged cells, forward-filled)
    Row 1: candidate names with party tags like "(D)", "(R)"
    Data starts at data_start_row (skipping Total/Percentage rows).
    """
    header_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    row0 = list(header_rows[0])
    row1 = list(header_rows[1])
    n = max(len(row0), len(row1))
    row0 += [None] * (n - len(row0))
    row1 += [None] * (n - len(row1))

    # Forward-fill race names
    col_race: list[str | None] = []
    cur = None
    for v in row0:
        if v is not None:
            cur = re.sub(r"\s+", " ", str(v)).strip()
        col_race.append(cur)

    # Map columns to (race_label, party)
    col_mapping: list[tuple[str, str] | None] = [None] * n

    for i in range(meta_cols, n):
        race_raw = col_race[i] if i < len(col_race) else None
        cand = row1[i] if i < len(row1) else None
        if not race_raw or not cand:
            continue

        cand_str = str(cand).strip()

        # Extract party from candidate name
        party = _extract_party_tag(cand_str)
        if party not in ("D", "R"):
            continue

        # Match race to our office map
        race_lower = race_raw.lower()
        race_label = None
        for kw, lbl in office_map.items():
            if kw in race_lower:
                race_label = lbl
                break

        if race_label:
            col_mapping[i] = (race_label, "d" if party == "D" else "r")

    # Read data rows
    records = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row[0] is None:
            continue
        county_name = str(row[0]).strip().upper()
        if county_name in ("", "TOTAL", "PERCENTAGE", "COUNTY", "COUNTY NAME"):
            continue

        for i, mapping in enumerate(col_mapping):
            if mapping is None or i >= len(row):
                continue
            race_label, party = mapping
            votes = row[i] if row[i] is not None else 0
            try:
                votes = float(votes)
            except (ValueError, TypeError):
                votes = 0
            records.append({
                "county_name": county_name,
                "race": race_label,
                "party": party,
                "votes": votes,
            })

    df = pd.DataFrame(records)
    if df.empty:
        return pd.DataFrame(columns=["county_name", "race", "d_votes", "r_votes"])

    # Pivot to county_name, race, d_votes, r_votes
    agg = df.groupby(["county_name", "race", "party"])["votes"].sum().reset_index()
    d_votes = agg[agg["party"] == "d"].rename(columns={"votes": "d_votes"}).drop(columns=["party"])
    r_votes = agg[agg["party"] == "r"].rename(columns={"votes": "r_votes"}).drop(columns=["party"])
    result = d_votes.merge(r_votes, on=["county_name", "race"], how="outer").fillna(0)

    return result


def _extract_party_tag(name: str) -> str:
    """Extract party from candidate name like 'Barack Obama (D)' → 'D'."""
    m = re.search(r"\(([A-Z])\)\s*$", name)
    return m.group(1) if m else "O"


# ---------------------------------------------------------------------------
# Existing SOS files (2016–2024) — aggregate to county level
# ---------------------------------------------------------------------------

def parse_existing_sos_county_votes(path: str | Path, year: str) -> pd.DataFrame:
    """
    Parse a 2016+ SOS file using the existing parser and aggregate to county level.

    Returns DataFrame: county_name, race, d_votes, r_votes
    """
    from src.ingest_sos import load_sos_file, COUNTY_COL

    sos = load_sos_file(path)

    records = []
    for label, spec in sos.statewide.items():
        if not spec.has_contest():
            continue

        df = sos.precinct_statewide
        d = df[spec.d_cols].sum(axis=1)
        r = df[spec.r_cols].sum(axis=1)

        county_agg = pd.DataFrame({
            COUNTY_COL: df[COUNTY_COL],
            "d_votes": d,
            "r_votes": r,
        }).groupby(COUNTY_COL)[["d_votes", "r_votes"]].sum().reset_index()

        county_agg["race"] = label
        county_agg = county_agg.rename(columns={COUNTY_COL: "county_name"})
        records.append(county_agg)

    if not records:
        return pd.DataFrame(columns=["county_name", "race", "d_votes", "r_votes"])

    result = pd.concat(records, ignore_index=True)
    _print_summary(result, year)
    return result


# ---------------------------------------------------------------------------
# Summary printer
# ---------------------------------------------------------------------------

def _print_summary(df: pd.DataFrame, year: str) -> None:
    """Print county vote totals summary for verification."""
    if df.empty:
        print(f"  {year}: No races parsed")
        return

    races = df["race"].unique()
    print(f"  {year}: {len(races)} races, {df['county_name'].nunique()} counties")
    for race in sorted(races):
        race_df = df[df["race"] == race]
        d_total = race_df["d_votes"].sum()
        r_total = race_df["r_votes"].sum()
        d_share = d_total / (d_total + r_total) if (d_total + r_total) > 0 else 0
        print(f"    {race:10s}  D={d_total:>10,.0f}  R={r_total:>10,.0f}  D-2P={d_share:.4f}")


# ---------------------------------------------------------------------------
# Unified county FIPS mapping
# ---------------------------------------------------------------------------

def county_name_to_fips(county_votes: pd.DataFrame) -> pd.DataFrame:
    """
    Add county_fips column to county_votes DataFrame.

    Maps county_name (uppercase) to 3-digit FIPS via the authoritative Ohio table.
    """
    from src.join_sos_vest import OHIO_FIPS_TO_COUNTY

    # Reverse map: county_name → FIPS
    name_to_fips = {name: fips for fips, name in OHIO_FIPS_TO_COUNTY.items()}

    county_votes = county_votes.copy()
    county_votes["county_fips"] = county_votes["county_name"].map(name_to_fips)

    unmapped = county_votes["county_fips"].isna().sum()
    if unmapped > 0:
        bad = county_votes[county_votes["county_fips"].isna()]["county_name"].unique()
        print(f"  Warning: {unmapped} rows with unmapped county names: {bad[:5]}")

    return county_votes.dropna(subset=["county_fips"])
